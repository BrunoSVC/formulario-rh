from flask import Flask, render_template, request
import os
from openpyxl import Workbook, load_workbook
from itsdangerous import URLSafeTimedSerializer

# 1️⃣ Criar app
app = Flask(__name__)

# 2️⃣ Configurações
UPLOAD_FOLDER = "uploads"
ARQUIVO_EXCEL = "candidatos.xlsx"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 3️⃣ Segurança (token)
secret = URLSafeTimedSerializer("chave_super_secreta")

def gerar_link():
    token = secret.dumps("acesso_formulario")

    base_url = request.host_url  # pega automaticamente a URL correta

    link = f"{base_url}formulario/{token}"

    return link

# 4️⃣ Rotas

@app.route("/")
def home():
    return "Sistema rodando! Use link com token."

@app.route("/formulario/<token>")
def formulario_token(token):
    try:
        secret.loads(token, max_age=3600)
    except:
        return "Link expirado ou inválido"

    return render_template("formulario.html")


@app.route("/enviar", methods=["POST"])
def enviar():

    nome = request.form["nome"]
    email = request.form["email"]
    telefone = request.form["telefone"]

    curriculo = request.files["curriculo"]
    documento = request.files["documento"]

    curriculo_nome = curriculo.filename
    documento_nome = documento.filename

    curriculo.save(os.path.join(UPLOAD_FOLDER, curriculo_nome))
    documento.save(os.path.join(UPLOAD_FOLDER, documento_nome))

    try:
        wb = load_workbook(ARQUIVO_EXCEL)
        planilha = wb.active
    except:
        wb = Workbook()
        planilha = wb.active
        planilha.append(["Nome", "Email", "Telefone", "Curriculo", "Documento"])

    planilha.append([nome, email, telefone, curriculo_nome, documento_nome])

    wb.save(ARQUIVO_EXCEL)

    return "Cadastro enviado com sucesso!"


@app.route("/admin")
def admin():

    from openpyxl import load_workbook

    # gerar link
    link = gerar_link()

    try:
        wb = load_workbook(ARQUIVO_EXCEL)
        planilha = wb.active
    except:
        return render_template("admin.html", dados=[], link=link)

    dados = []

    for linha in planilha.iter_rows(min_row=2, values_only=True):
        dados.append(linha)

    return render_template("admin.html", dados=dados, link=link)
   

# 5️⃣ Rodar app (SEMPRE POR ÚLTIMO)
if __name__ == "__main__":
    app.run(debug=True)